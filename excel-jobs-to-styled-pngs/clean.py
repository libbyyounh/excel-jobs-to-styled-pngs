#!/usr/bin/env python3
"""Stage 0 — Clean a recruitment Excel into a tidy format that render.py can consume.

Input : xlsx path (CLI arg)
Output: <stem>_cleaned.xlsx       — one row per (company, job), no section/policy noise
        <stem>_clean_report.txt   — list of dropped rows and the reason for each

Configuration:
- CONFIGS = []  → auto-detect header row + column indices for every sheet in the workbook
- CONFIGS = [(sheet, min_row, comp_col, job_col, sal_col), ...]  → manual override

Auto-detection scans the first 5 rows for a cell whose value is '序号' (= header row), then
identifies the company / job / salary columns by matching common header strings. Sheets where
auto-detection fails are skipped with a warning; if you hit one, add a manual entry to CONFIGS.
"""
import openpyxl
import sys
from pathlib import Path

# Manual override. Leave as [] to auto-detect every sheet in the workbook.
# Format: (sheet_name, min_data_row, company_col, job_col, salary_col)
CONFIGS = []

# Header strings we recognise. Add to these if your file uses a different label.
COMPANY_HEADERS = ('公司名称', '所属公司', '公司', '招聘单位', '企业名称')
JOB_HEADERS     = ('岗位名称', '岗位', '职位', '招聘岗位')
SALARY_HEADERS  = ('薪资范围', '薪酬范围', '薪酬', '薪资', '工资', '月薪')

# 公司列里出现这些关键字的行 → 当政策/资料说明丢掉
NOISE_KEYWORDS = ('政策', '贷款', '详见')

# 序号列以这些前缀开头的行 → section header
SECTION_PREFIXES_SKIP  = ('一、',)   # 单条 skip
SECTION_PREFIXES_BREAK = ('二、',)   # break 整个 sheet 的处理


def auto_detect(ws):
    """Scan first 5 rows for a header row containing '序号', then locate the data columns.

    Returns (header_row, company_col, job_col, salary_col) on success, or None if any column
    is missing. `header_row` is 1-based (matches openpyxl row numbers).
    """
    for header_row in range(1, 6):
        cells = [c for c in next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))]
        if '序号' not in [str(c).strip() for c in cells if c is not None]:
            continue
        company_col = job_col = salary_col = None
        for idx, cell in enumerate(cells):
            if cell is None:
                continue
            s = str(cell).strip()
            if company_col is None and s in COMPANY_HEADERS:
                company_col = idx
            elif job_col is None and s in JOB_HEADERS:
                job_col = idx
            elif salary_col is None and s in SALARY_HEADERS:
                salary_col = idx
        if all(x is not None for x in (company_col, job_col, salary_col)):
            return (header_row, company_col, job_col, salary_col)
    return None


def iter_clean(ws, min_row, company_col, job_col, salary_col):
    """Yield (excel_row, drop_reason | None, record | None) for every input row.

    `drop_reason` is a string explaining why the row was dropped (or None if kept).
    `record` is a dict {company, job, salary} when the row is kept (else None).
    Stops iterating when a "二、" section header is hit.
    """
    current_company = None
    for excel_row, row in enumerate(
        ws.iter_rows(min_row=min_row, values_only=True), start=min_row
    ):
        seq = row[1] if len(row) > 1 and row[1] else ''

        if isinstance(seq, str):
            if any(seq.startswith(p) for p in SECTION_PREFIXES_BREAK):
                return
            if any(seq.startswith(p) for p in SECTION_PREFIXES_SKIP):
                yield excel_row, 'section_header', None
                continue
            if seq == '序号':
                yield excel_row, 'header', None
                continue

        company = (row[company_col] or '').strip() if len(row) > company_col else ''
        job     = (row[job_col]     or '').strip() if len(row) > job_col     else ''
        salary  = (row[salary_col]  or '').strip() if len(row) > salary_col  else ''

        if not company and not job and not salary:
            yield excel_row, 'empty', None
            continue
        if not job or not salary:
            yield excel_row, 'missing_data', None
            continue
        if any(kw in company for kw in NOISE_KEYWORDS):
            yield excel_row, f'noise({company})', None
            continue

        if company:
            current_company = company
        if not current_company:
            yield excel_row, 'no_company_yet', None
            continue

        yield excel_row, None, {
            'company': current_company.replace('\n', ' '),
            'job':     job,
            'salary':  salary,
        }


def clean(xlsx_path):
    xlsx_path = Path(xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    cleaned_wb = openpyxl.Workbook()
    cleaned_wb.remove(cleaned_wb.active)
    report = [f'Clean report for {xlsx_path.name}', '']

    # Build the work list: manual CONFIGS if any, else auto-detect every sheet.
    if CONFIGS:
        work = list(CONFIGS)
    else:
        work = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            detected = auto_detect(ws)
            if detected is None:
                report.append(f'--- [{sheet_name}] ---')
                report.append('  SKIPPED: could not auto-detect header (no 序号 cell, or missing 公司/岗位/薪酬 column).')
                report.append('  Fix: add a manual entry to CONFIGS in clean.py.')
                continue
            header_row, comp_c, job_c, sal_c = detected
            work.append((sheet_name, header_row + 1, comp_c, job_c, sal_c))
            report.append(f'--- [{sheet_name}] ---')
            report.append(f'  auto-detected: header_row={header_row}  comp={comp_c}  job={job_c}  salary={sal_c}')

    for sheet_name, min_row, comp_c, job_c, sal_c in work:
        if sheet_name not in wb.sheetnames:
            report.append(f'--- [{sheet_name}] ---')
            report.append('  NOT FOUND, skipped')
            continue
        ws = wb[sheet_name]
        out_ws = cleaned_wb.create_sheet(sheet_name)
        out_ws.append(['序号', '公司', '岗位', '薪酬'])

        kept, dropped = 0, 0
        idx = 0
        for excel_row, drop_reason, record in iter_clean(ws, min_row, comp_c, job_c, sal_c):
            if drop_reason:
                report.append(f'  drop A{excel_row}: {drop_reason}')
                dropped += 1
            else:
                idx += 1
                out_ws.append([idx, record['company'], record['job'], record['salary']])
                kept += 1
        if sheet_name in [s for s, *_ in CONFIGS] or CONFIGS:
            report.append(f'  kept={kept}  dropped={dropped}')
        else:
            report.append(f'  kept={kept}  dropped={dropped}')

    cleaned_path = xlsx_path.with_name(f'{xlsx_path.stem}_cleaned.xlsx')
    cleaned_wb.save(cleaned_path)
    report_path = xlsx_path.with_name(f'{xlsx_path.stem}_clean_report.txt')
    report_path.write_text('\n'.join(report) + '\n', encoding='utf-8')
    print(f'cleaned → {cleaned_path}')
    print(f'report  → {report_path}')
    return cleaned_path


if __name__ == '__main__':
    if len(sys.argv) < 2:
        raise SystemExit('usage: clean.py <recruitment.xlsx>')
    clean(sys.argv[1])
