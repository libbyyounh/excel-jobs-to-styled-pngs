#!/usr/bin/env python3
"""Stage 1 — Render a *cleaned* recruitment Excel into a designed long-image PNG.

Input : <stem>_cleaned.xlsx  (output of clean.py — already filtered, no section headers,
        no policy noise, no empty rows; columns are exactly 序号/公司/岗位/薪酬 in that order)
Output: one or more PNG files

If you have a *raw* Excel, run `clean.py` first. This script does no data filtering.

To use a different style, change the constants in the "Visual spec" section below.
"""
import openpyxl
import re
import sys
from pathlib import Path
from PIL import Image, ImageDraw, ImageFont

# (sheet_name, output_png). Leave as [] to render every sheet in the workbook with an
# auto-derived output filename (c_{sanitized_sheet_name}.png).
TARGETS = []


def safe_filename(sheet_name):
    """Sanitize a sheet name for use as a PNG filename."""
    return re.sub(r'[\\/:*?"<>|.\s]+', '_', sheet_name).strip('_') or 'sheet'

# --- Visual spec (default teal-badge card) ---
# Override these if the user supplies a different style reference.
WIDTH          = 1200
MARGIN_LEFT    = 60
MARGIN_RIGHT   = 60
BG_COLOR       = (255, 255, 255)
TEAL           = (100, 149, 137)
DARK_TEXT      = (51, 51, 51)
BADGE_W        = 180
BADGE_H        = 56
COMPANY_LINE_H = 42
JOB_ROW_H      = 58
SECTION_GAP    = 30
SEPARATOR_COL  = (200, 200, 200)

# CJK font fallback chain (probed in order; first that loads wins)
FONT_CANDIDATES = [
    '/System/Library/Fonts/STHeiti Medium.ttc',
    '/System/Library/Fonts/Hiragino Sans GB.ttc',
    '/Library/Fonts/PingFang.ttc',
    '/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc',
    '/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc',
]
FONT_SIZES = {'badge': 38, 'company': 40, 'job': 32, 'salary': 30}

# Cleaned xlsx always has these columns in this order
COL_SEQ, COL_COMPANY, COL_JOB, COL_SALARY = 0, 1, 2, 3


def load_fonts():
    for path in FONT_CANDIDATES:
        try:
            ImageFont.truetype(path, 12)
            return {name: ImageFont.truetype(path, size) for name, size in FONT_SIZES.items()}
        except OSError:
            continue
    raise RuntimeError(f'No usable CJK font found in any of: {FONT_CANDIDATES}')


def wrap_text(text, font, max_w, tmp_img):
    draw = ImageDraw.Draw(tmp_img)
    lines, current = [], ''
    for ch in text:
        test = current + ch
        if draw.textbbox((0, 0), test, font=font)[2] > max_w:
            if current:
                lines.append(current)
            current = ch
        else:
            current = test
    if current:
        lines.append(current)
    return lines or ['']


def load_jobs(ws):
    """Read cleaned sheet (序号/公司/岗位/薪酬) → [(company, [jobs...])] preserving order."""
    companies = []
    current = None
    for row in ws.iter_rows(min_row=2, values_only=True):  # row 1 is the header
        company = (row[COL_COMPANY] or '').strip()
        job     = (row[COL_JOB]     or '').strip()
        salary  = (row[COL_SALARY]  or '').strip()
        if not company or not job:
            continue  # cleaned file shouldn't have these, but guard anyway
        if company != current:
            current = company
            companies.append([company, []])
        companies[-1][1].append({'job': job, 'salary': salary})
    return companies


def render(sheet_name, output_png, fonts, xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        print(f'  skip {sheet_name}: not in {xlsx_path.name}')
        return
    companies = load_jobs(wb[sheet_name])
    if not companies:
        print(f'  skip {sheet_name}: no jobs after clean')
        return

    company_max_w = WIDTH - MARGIN_LEFT - MARGIN_RIGHT - BADGE_W - 24
    tmp_img = Image.new('RGB', (1, 1))
    wrapped = [
        (name, jbs, wrap_text(name, fonts['company'], company_max_w, tmp_img))
        for name, jbs in companies
    ]

    # Card header height = max(badge, wrapped company lines) — side by side, not stacked
    total_h = 40
    for _, jbs, lines in wrapped:
        total_h += max(BADGE_H, COMPANY_LINE_H * len(lines))
        total_h += len(jbs) * JOB_ROW_H
        total_h += SECTION_GAP
    total_h += 40

    img = Image.new('RGB', (WIDTH, total_h), BG_COLOR)
    draw = ImageDraw.Draw(img)
    y = 40

    for idx, (company_name, jbs, lines) in enumerate(wrapped):
        header_h = max(BADGE_H, COMPANY_LINE_H * len(lines))
        badge_x, badge_y = MARGIN_LEFT, y

        draw.rounded_rectangle(
            [badge_x, badge_y, badge_x + BADGE_W, badge_y + BADGE_H],
            radius=12, fill=TEAL,
        )
        draw.text(
            (badge_x + 16, badge_y + 6),
            '招聘岗位', fill=(255, 255, 255), font=fonts['badge'],
        )

        company_x = badge_x + BADGE_W + 24
        for line_idx, line in enumerate(lines):
            draw.text(
                (company_x, y + 6 + line_idx * COMPANY_LINE_H),
                line, fill=DARK_TEXT, font=fonts['company'],
            )
        y += header_h

        for job in jbs:
            job_x = MARGIN_LEFT + 30
            draw.text((job_x, y + 10), job['job'], fill=DARK_TEXT, font=fonts['job'])
            # Salary goes right after the job name, with a 40 px gap — NOT a fixed column
            job_bbox = draw.textbbox((job_x, y + 10), job['job'], font=fonts['job'])
            draw.text(
                (job_bbox[2] + 40, y + 12),
                job['salary'], fill=DARK_TEXT, font=fonts['salary'],
            )
            y += JOB_ROW_H

        if idx < len(wrapped) - 1:
            y += 10
            draw.line(
                [(MARGIN_LEFT, y), (WIDTH - MARGIN_RIGHT, y)],
                fill=SEPARATOR_COL, width=2,
            )
            y += SECTION_GAP - 10

    img.save(output_png, 'PNG', quality=95)
    print(f'  {sheet_name} → {output_png}  size={img.size}  '
          f'companies={len(companies)}  jobs={sum(len(j) for _, j in companies)}')


def main():
    if len(sys.argv) < 2:
        raise SystemExit('usage: render.py <cleaned.xlsx>')
    xlsx = Path(sys.argv[1])
    if not xlsx.exists():
        raise SystemExit(f'Cleaned xlsx not found: {xlsx}\nRun clean.py first.')
    fonts = load_fonts()
    if TARGETS:
        work = list(TARGETS)
    else:
        wb = openpyxl.load_workbook(xlsx, data_only=True)
        work = [(name, f'c_{safe_filename(name)}.png') for name in wb.sheetnames]
    for sheet_name, out_png in work:
        render(sheet_name, out_png, fonts, xlsx)


if __name__ == '__main__':
    main()
